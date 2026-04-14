from __future__ import annotations

import csv
import json
import mimetypes
import re
from datetime import datetime
from email.parser import BytesParser
from email.policy import default as email_policy
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from app.backends import validate_docker_backend_access
from app.chat_history_store import (
    delete_thread_history,
    ensure_chat_history_schema,
    fetch_latest_thread_history,
    fetch_thread_history,
    fetch_thread_ui_state,
    list_history_threads,
    upsert_thread_ui_state,
    upsert_chat_session,
)
from app.config import load_settings
from app.demo_session import run_demo_session_stream
from app.prompts import get_prompt_sections, reset_prompt_section, update_prompt_section
from app.workspace_files import WORKSPACE_ROOT, build_workspace_file_card, resolve_workspace_file, write_workspace_text_file


PROJECT_ROOT = Path(__file__).resolve().parent.parent
FRONTEND_ROOT = PROJECT_ROOT / "frontend_demo"
ALLOWED_USER_FILE_EXTENSIONS = {".md", ".xlsx", ".csv", ".txt", ".py"}
MAX_USER_FILE_SIZE = 10 * 1024
MAX_USER_FILE_COUNT = 3


class DemoRequestHandler(SimpleHTTPRequestHandler):
    protocol_version = "HTTP/1.1"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(FRONTEND_ROOT), **kwargs)

    def do_POST(self) -> None:
        if self.path == "/api/demo/prompts":
            self._handle_update_prompt()
            return
        if self.path == "/api/demo/prompts/reset":
            self._handle_reset_prompt()
            return
        if self.path == "/api/demo/thread-state":
            self._handle_update_thread_state()
            return
        if self.path == "/api/demo/workspace-file":
            self._handle_update_workspace_file()
            return
        if self.path == "/api/demo/user-file":
            self._handle_upload_user_file()
            return
        if self.path != "/api/demo/run":
            self.send_error(HTTPStatus.NOT_FOUND, "Unknown API endpoint")
            return

        content_length = int(self.headers.get("Content-Length", "0"))
        raw_body = self.rfile.read(content_length) if content_length else b"{}"
        content_type = (self.headers.get("Content-Type") or "").strip().lower()
        uploaded_user_files: list[dict[str, object]] = []
        if content_type.startswith("multipart/form-data"):
            try:
                payload, uploaded_user_files = _parse_multipart_form_data(self.headers.get("Content-Type", ""), raw_body)
            except ValueError as exc:
                self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
                return
        else:
            try:
                payload = json.loads(raw_body.decode("utf-8"))
            except json.JSONDecodeError:
                self._send_json({"error": "invalid_json"}, status=HTTPStatus.BAD_REQUEST)
                return

        query = str(payload.get("query", "")).strip()
        thread_id = str(payload.get("thread_id", "")).strip()
        session_id = str(payload.get("session_id", "")).strip()
        max_rounds = int(payload.get("max_rounds", 12) or 12)
        raw_messages = payload.get("messages", [])
        if not query:
            self._send_json({"error": "query_required"}, status=HTTPStatus.BAD_REQUEST)
            return
        if not thread_id:
            self._send_json({"error": "thread_id_required"}, status=HTTPStatus.BAD_REQUEST)
            return
        if not session_id:
            self._send_json({"error": "session_id_required"}, status=HTTPStatus.BAD_REQUEST)
            return

        messages = []
        if isinstance(raw_messages, list):
            for item in raw_messages:
                if not isinstance(item, dict):
                    continue
                role = str(item.get("role", "")).strip()
                content = str(item.get("content", "")).strip()
                if role in {"user", "assistant", "system"} and content:
                    messages.append({"role": role, "content": content})

        try:
            user_files = _normalize_user_file_refs(payload.get("user_files", []), thread_id)
            user_files.extend(_save_uploaded_user_file(item, thread_id) for item in uploaded_user_files)
        except ValueError as exc:
            self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        agent_query = _compose_agent_query(query, user_files)
        if messages:
            patched_messages = messages[:-1]
            last_message = messages[-1]
            if last_message.get("role") == "user":
                patched_messages.append({"role": "user", "content": agent_query})
            else:
                patched_messages = [*messages, {"role": "user", "content": agent_query}]
            messages = patched_messages
        else:
            messages = [{"role": "user", "content": agent_query}]

        try:
            settings = load_settings(argv=[])
            self._send_ndjson_stream(
                settings=settings,
                query=query,
                agent_query=agent_query,
                thread_id=thread_id,
                session_id=session_id,
                max_rounds=max_rounds,
                messages=messages,
                user_files=user_files,
            )
        except Exception as exc:  # noqa: BLE001
            self._send_json({"error": str(exc)}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
            return

    def _handle_update_prompt(self) -> None:
        content_length = int(self.headers.get("Content-Length", "0"))
        raw_body = self.rfile.read(content_length) if content_length else b"{}"
        try:
            payload = json.loads(raw_body.decode("utf-8"))
        except json.JSONDecodeError:
            self._send_json({"error": "invalid_json"}, status=HTTPStatus.BAD_REQUEST)
            return

        prompt_id = str(payload.get("id", "")).strip()
        content = str(payload.get("content", ""))
        if not prompt_id:
            self._send_json({"error": "prompt_id_required"}, status=HTTPStatus.BAD_REQUEST)
            return
        if not content.strip():
            self._send_json({"error": "prompt_content_required"}, status=HTTPStatus.BAD_REQUEST)
            return

        try:
            updated = update_prompt_section(prompt_id=prompt_id, content=content)
        except KeyError as exc:
            self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return
        except ValueError as exc:
            self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        self._send_json({"prompt": updated, "prompts": get_prompt_sections(max_rounds=12)}, status=HTTPStatus.OK)

    def _handle_reset_prompt(self) -> None:
        content_length = int(self.headers.get("Content-Length", "0"))
        raw_body = self.rfile.read(content_length) if content_length else b"{}"
        try:
            payload = json.loads(raw_body.decode("utf-8"))
        except json.JSONDecodeError:
            self._send_json({"error": "invalid_json"}, status=HTTPStatus.BAD_REQUEST)
            return

        prompt_id = str(payload.get("id", "")).strip()
        if not prompt_id:
            self._send_json({"error": "prompt_id_required"}, status=HTTPStatus.BAD_REQUEST)
            return

        try:
            reset_prompt = reset_prompt_section(prompt_id=prompt_id)
        except KeyError as exc:
            self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        self._send_json({"prompt": reset_prompt, "prompts": get_prompt_sections(max_rounds=12)}, status=HTTPStatus.OK)

    def _handle_update_thread_state(self) -> None:
        content_length = int(self.headers.get("Content-Length", "0"))
        raw_body = self.rfile.read(content_length) if content_length else b"{}"
        try:
            payload = json.loads(raw_body.decode("utf-8"))
        except json.JSONDecodeError:
            self._send_json({"error": "invalid_json"}, status=HTTPStatus.BAD_REQUEST)
            return

        thread_id = str(payload.get("thread_id", "")).strip()
        ui_state = payload.get("ui_state", {})
        if not thread_id:
            self._send_json({"error": "thread_id_required"}, status=HTTPStatus.BAD_REQUEST)
            return
        if not isinstance(ui_state, dict):
            self._send_json({"error": "ui_state_must_be_object"}, status=HTTPStatus.BAD_REQUEST)
            return

        settings = load_settings(argv=[])
        upsert_thread_ui_state(settings=settings, thread_id=thread_id, ui_state=ui_state)
        self._send_json({"ok": True})

    def _handle_update_workspace_file(self) -> None:
        content_length = int(self.headers.get("Content-Length", "0"))
        raw_body = self.rfile.read(content_length) if content_length else b"{}"
        try:
            payload = json.loads(raw_body.decode("utf-8"))
        except json.JSONDecodeError:
            self._send_json({"error": "invalid_json"}, status=HTTPStatus.BAD_REQUEST)
            return

        relative_path = str(payload.get("path", "")).strip()
        content = payload.get("content")
        if not relative_path:
            self._send_json({"error": "path_required"}, status=HTTPStatus.BAD_REQUEST)
            return
        if not isinstance(content, str):
            self._send_json({"error": "content_must_be_string"}, status=HTTPStatus.BAD_REQUEST)
            return

        try:
            file_payload = write_workspace_text_file(relative_path, content)
        except FileNotFoundError:
            self._send_json({"error": "file_not_found"}, status=HTTPStatus.NOT_FOUND)
            return
        except ValueError as exc:
            self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        self._send_json({"ok": True, "file": file_payload}, status=HTTPStatus.OK)

    def _handle_upload_user_file(self) -> None:
        content_length = int(self.headers.get("Content-Length", "0"))
        raw_body = self.rfile.read(content_length) if content_length else b""
        content_type = self.headers.get("Content-Type", "")
        try:
            payload, uploaded_user_files = _parse_multipart_form_data(content_type, raw_body)
        except ValueError as exc:
            self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        thread_id = str(payload.get("thread_id", "")).strip()
        if not thread_id:
            self._send_json({"error": "thread_id_required"}, status=HTTPStatus.BAD_REQUEST)
            return
        if len(uploaded_user_files) != 1:
            self._send_json({"error": "一次只能上传一个文件。"}, status=HTTPStatus.BAD_REQUEST)
            return

        try:
            file_payload = _save_uploaded_user_file(uploaded_user_files[0], thread_id, finalized=False)
        except ValueError as exc:
            self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        self._send_json({"ok": True, "file": file_payload}, status=HTTPStatus.OK)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if self.path == "/api/health":
            self._send_json({"ok": True})
            return
        if self.path == "/api/demo/meta":
            settings = load_settings(argv=[])
            self._send_json({"agents": [], "model": settings.model})
            return
        if self.path == "/api/demo/prompts":
            settings = load_settings(argv=[])
            self._send_json({"prompts": get_prompt_sections(max_rounds=12), "model": settings.model})
            return
        if parsed.path == "/api/demo/history":
            settings = load_settings(argv=[])
            params = parse_qs(parsed.query)
            thread_id = (params.get("thread_id") or [""])[0].strip()
            history = (
                fetch_thread_history(settings=settings, thread_id=thread_id)
                if thread_id
                else fetch_latest_thread_history(settings=settings)
            )
            if history:
                history["ui_state"] = fetch_thread_ui_state(settings=settings, thread_id=history["thread_id"]) or {}
            self._send_json(history or {"thread_id": "", "sessions": [], "ui_state": {}})
            return
        if parsed.path == "/api/demo/history/threads":
            settings = load_settings(argv=[])
            self._send_json({"threads": list_history_threads(settings)})
            return
        if parsed.path == "/api/demo/thread-state":
            settings = load_settings(argv=[])
            params = parse_qs(parsed.query)
            thread_id = (params.get("thread_id") or [""])[0].strip()
            self._send_json({"thread_id": thread_id, "ui_state": fetch_thread_ui_state(settings=settings, thread_id=thread_id) or {}})
            return
        if parsed.path == "/api/demo/workspace-file":
            self._handle_workspace_file(parsed)
            return
        if self.path == "/":
            self.path = "/index.html"
        super().do_GET()

    def do_DELETE(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/demo/user-file":
            self._handle_delete_user_file(parsed)
            return
        if parsed.path != "/api/demo/history":
            self.send_error(HTTPStatus.NOT_FOUND, "Unknown API endpoint")
            return

        settings = load_settings(argv=[])
        params = parse_qs(parsed.query)
        thread_id = (params.get("thread_id") or [""])[0].strip()
        if not thread_id:
            self._send_json({"error": "thread_id_required"}, status=HTTPStatus.BAD_REQUEST)
            return

        deleted = delete_thread_history(settings=settings, thread_id=thread_id)
        latest = fetch_latest_thread_history(settings=settings)
        self._send_json(
            {
                "ok": deleted,
                "deleted_thread_id": thread_id,
                "latest_thread_id": latest["thread_id"] if latest else "",
            },
            status=HTTPStatus.OK if deleted else HTTPStatus.NOT_FOUND,
        )

    def _handle_delete_user_file(self, parsed) -> None:
        params = parse_qs(parsed.query)
        relative_path = (params.get("path") or [""])[0].strip()
        if not relative_path:
            self._send_json({"error": "path_required"}, status=HTTPStatus.BAD_REQUEST)
            return
        if not relative_path.startswith("user_file/"):
            self._send_json({"error": "只能删除 user_file 下的文件。"}, status=HTTPStatus.BAD_REQUEST)
            return
        try:
            file_path = resolve_workspace_file(relative_path)
        except FileNotFoundError:
            self._send_json({"ok": True, "deleted": False}, status=HTTPStatus.OK)
            return
        except ValueError as exc:
            self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return
        file_path.unlink(missing_ok=True)
        self._send_json({"ok": True, "deleted": True, "path": relative_path}, status=HTTPStatus.OK)

    def end_headers(self) -> None:
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def log_message(self, format: str, *args) -> None:
        return

    def _send_json(self, payload: dict, status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_ndjson_stream(
        self,
        settings,
        query: str,
        agent_query: str,
        thread_id: str,
        session_id: str,
        max_rounds: int,
        messages: list[dict[str, str]] | None = None,
        user_files: list[dict] | None = None,
    ) -> None:
        self.close_connection = True
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/x-ndjson; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Connection", "close")
        self.end_headers()

        for event in run_demo_session_stream(
            settings=settings,
            query=query,
            agent_query=agent_query,
            max_rounds=max_rounds,
            messages=messages,
            user_files=user_files,
        ):
            try:
                line = json.dumps(event, ensure_ascii=False).encode("utf-8") + b"\n"
                self.wfile.write(line)
                self.wfile.flush()
                upsert_chat_session(
                    settings=settings,
                    thread_id=thread_id,
                    session_id=session_id,
                    query_text=query,
                    payload=event["payload"],
                    error_text=event["payload"].get("stop_reason", "") if event["type"] == "error" else "",
                )
            except (BrokenPipeError, ConnectionResetError):
                break

    def _handle_workspace_file(self, parsed) -> None:
        params = parse_qs(parsed.query)
        relative_path = (params.get("path") or [""])[0]
        download = ((params.get("download") or ["0"])[0] == "1")
        preview_format = ((params.get("format") or [""])[0] or "").strip().lower()
        if not relative_path.strip():
            self._send_json({"error": "path_required"}, status=HTTPStatus.BAD_REQUEST)
            return
        try:
            file_path = resolve_workspace_file(relative_path)
        except FileNotFoundError:
            self._send_json({"error": "file_not_found"}, status=HTTPStatus.NOT_FOUND)
            return
        except ValueError as exc:
            self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        if preview_format == "json":
            if file_path.suffix.lower() not in {".xlsx", ".xls", ".csv"}:
                self._send_json({"error": "json_preview_not_supported"}, status=HTTPStatus.BAD_REQUEST)
                return
            try:
                payload = _build_spreadsheet_preview(file_path)
            except RuntimeError as exc:
                self._send_json({"error": str(exc)}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return
            except Exception as exc:  # noqa: BLE001
                self._send_json({"error": str(exc)}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return
            self._send_json(payload, status=HTTPStatus.OK)
            return

        body = file_path.read_bytes()
        mime_type, _ = mimetypes.guess_type(file_path.name)
        mime_type = mime_type or "application/octet-stream"
        disposition = "attachment" if download else "inline"
        download_name = _ascii_download_name(file_path.name)
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", mime_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Content-Disposition", f'{disposition}; filename="{download_name}"')
        self.end_headers()
        self.wfile.write(body)


def _parse_multipart_form_data(content_type: str, raw_body: bytes) -> tuple[dict[str, object], list[dict[str, object]]]:
    if "boundary=" not in content_type:
        raise ValueError("missing_multipart_boundary")

    parser = BytesParser(policy=email_policy)
    message = parser.parsebytes(
        f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8") + raw_body
    )
    payload: dict[str, object] = {}
    uploaded_files: list[dict[str, object]] = []
    if not message.is_multipart():
        raise ValueError("invalid_multipart_payload")

    for part in message.iter_parts():
        name = part.get_param("name", header="content-disposition")
        if not name:
            continue
        filename = part.get_filename()
        body = part.get_payload(decode=True) or b""
        if filename:
            uploaded_files.append(
                {
                "field_name": name,
                "filename": filename,
                "content_type": part.get_content_type() or "application/octet-stream",
                "content": body,
                }
            )
            continue
        payload[name] = body.decode(part.get_content_charset() or "utf-8").strip()

    if "messages" in payload:
        try:
            payload["messages"] = json.loads(str(payload["messages"]))
        except json.JSONDecodeError as exc:
            raise ValueError("invalid_messages_json") from exc
    if len(uploaded_files) > MAX_USER_FILE_COUNT:
        raise ValueError(f"最多上传 {MAX_USER_FILE_COUNT} 个文件。")
    return payload, uploaded_files


def _sanitize_filename_stem(name: str) -> str:
    stem = Path(name).stem.strip()
    normalized = re.sub(r"[^A-Za-z0-9._-]+", "_", stem)
    return normalized.strip("._") or "upload"


def _sanitize_token(value: str, *, fallback: str) -> str:
    normalized = re.sub(r"[^A-Za-z0-9_-]+", "_", (value or "").strip())
    return normalized.strip("._-") or fallback


def _build_user_file_name(filename: str, thread_id: str, *, finalized: bool) -> str:
    extension = Path(filename).suffix.lower()
    safe_thread_id = _sanitize_token(thread_id, fallback="thread")
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    stem = _sanitize_filename_stem(filename)
    if finalized:
        return f"{stem}_{safe_thread_id}_{timestamp}{extension}"
    return f"{stem}__pending__{safe_thread_id}_{timestamp}{extension}"


def _save_uploaded_user_file(uploaded_file: dict[str, object], thread_id: str, *, finalized: bool = False) -> dict[str, object]:
    filename = str(uploaded_file.get("filename", "")).strip()
    if not filename:
        raise ValueError("上传文件缺少文件名。")

    extension = Path(filename).suffix.lower()
    if extension not in ALLOWED_USER_FILE_EXTENSIONS:
        raise ValueError("仅支持 .md, .xlsx, .csv, .txt, .py 文件。")

    content = uploaded_file.get("content", b"")
    if not isinstance(content, bytes):
        raise ValueError("上传文件内容无效。")
    if len(content) > MAX_USER_FILE_SIZE:
        raise ValueError(f"文件不能超过 {MAX_USER_FILE_SIZE} bytes。")

    stored_name = _build_user_file_name(filename, thread_id, finalized=finalized)
    relative_path = f"user_file/{stored_name}"
    file_path = (WORKSPACE_ROOT / relative_path).resolve()
    file_path.parent.mkdir(parents=True, exist_ok=True)
    file_path.write_bytes(content)

    card = build_workspace_file_card(relative_path, title=filename)
    card["original_name"] = filename
    card["source"] = "user_upload"
    return card


def _normalize_user_file_refs(raw_user_files: object, thread_id: str) -> list[dict[str, object]]:
    if not raw_user_files:
        return []
    if not isinstance(raw_user_files, list):
        raise ValueError("user_files 必须是数组。")
    if len(raw_user_files) > MAX_USER_FILE_COUNT:
        raise ValueError(f"最多上传 {MAX_USER_FILE_COUNT} 个文件。")

    normalized: list[dict[str, object]] = []
    safe_thread_id = _sanitize_token(thread_id, fallback="thread")
    for item in raw_user_files:
        if not isinstance(item, dict):
            continue
        relative_path = str(item.get("path", "")).strip()
        if not relative_path:
            continue
        if not relative_path.startswith("user_file/"):
            raise ValueError("用户文件路径必须位于 user_file/ 下。")
        original_name = str(item.get("original_name") or item.get("name") or "").strip()
        file_path = resolve_workspace_file(relative_path)
        target_name = _build_user_file_name(original_name or file_path.name, thread_id, finalized=True)
        if "__pending__" in file_path.name and file_path.name != target_name:
            target_path = file_path.with_name(target_name)
            file_path.rename(target_path)
            relative_path = f"user_file/{target_name}"
        file_payload = build_workspace_file_card(relative_path, title=original_name or file_path.name)
        if safe_thread_id not in str(file_payload.get("name", "")):
            raise ValueError("用户文件与当前线程不匹配。")
        file_payload["original_name"] = original_name or file_payload["name"]
        file_payload["source"] = "user_upload"
        normalized.append(file_payload)
    return normalized


def _compose_agent_query(query: str, user_files: list[dict[str, object]]) -> str:
    if not user_files:
        return query

    file_lines = []
    for file in user_files:
        file_lines.append(
            "- 文件: {name}\n  格式: {ext}\n  路径: {path}".format(
                name=str(file.get("original_name") or file.get("name") or ""),
                ext=str(file.get("extension") or ""),
                path=str(file.get("path") or ""),
            )
        )
    return (
        f"{query}\n\n"
        "用户本轮还上传了以下文件。路径相对于当前 workspace 根目录，可直接读取：\n"
        f"{chr(10).join(file_lines)}"
    )


def _ascii_download_name(name: str) -> str:
    base = name.encode("ascii", errors="ignore").decode("ascii")
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", base).strip("._")
    return base or "download"


def _build_spreadsheet_preview(file_path: Path) -> dict[str, object]:
    if file_path.suffix.lower() == ".csv":
        return _build_csv_preview(file_path)

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel 预览依赖 openpyxl，请先安装 requirements。") from exc

    preview_row_limit = 40
    preview_col_limit = 20
    workbook = load_workbook(filename=file_path, read_only=False, data_only=True)
    sheets: list[dict[str, object]] = []
    try:
        for sheet in workbook.worksheets[:6]:
            max_row = sheet.max_row or 0
            max_col = sheet.max_column or 0
            limited_rows = min(max_row, preview_row_limit)
            limited_cols = min(max_col, preview_col_limit)
            merged_map = _build_merge_map(sheet, limited_rows, limited_cols)
            rows = _build_spreadsheet_rows(sheet, limited_rows, limited_cols, merged_map)
            column_widths = [
                _excel_width_to_px(getattr(sheet.column_dimensions.get(_column_letter(index + 1)), "width", None))
                for index in range(limited_cols)
            ]
            sheets.append(
                {
                    "name": sheet.title,
                    "row_count": max_row,
                    "column_count": max_col,
                    "preview_row_count": limited_rows,
                    "preview_column_count": limited_cols,
                    "truncated_rows": max_row > limited_rows,
                    "truncated_columns": max_col > limited_cols,
                    "column_widths": column_widths,
                    "rows": rows,
                }
            )
    finally:
        workbook.close()

    return {
        "kind": "spreadsheet",
        "name": file_path.name,
        "sheet_count": len(sheets),
        "preview_limits": {
            "max_rows": preview_row_limit,
            "max_columns": preview_col_limit,
        },
        "sheets": sheets,
    }


def _coerce_spreadsheet_cell(value) -> str:
    if value is None:
        return ""
    return str(value)


def _build_csv_preview(file_path: Path) -> dict[str, object]:
    preview_row_limit = 40
    preview_col_limit = 20
    try:
        raw_text = file_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        raw_text = file_path.read_text(encoding="utf-8-sig")

    reader = csv.reader(raw_text.splitlines())
    all_rows = [list(row) for row in reader]
    row_count = len(all_rows)
    column_count = max((len(row) for row in all_rows), default=0)
    limited_rows = min(row_count, preview_row_limit)
    limited_cols = min(column_count, preview_col_limit)
    rows: list[list[dict[str, object]]] = []
    column_widths = [96 for _ in range(limited_cols)]
    for row in all_rows[:limited_rows]:
        visible_row: list[dict[str, object]] = []
        for index in range(limited_cols):
            cell = row[index] if index < len(row) else ""
            text = _coerce_spreadsheet_cell(cell)
            column_widths[index] = max(column_widths[index], _infer_text_width_px(text))
            visible_row.append(
                {
                    "value": text,
                    "column": index,
                    "rowspan": 1,
                    "colspan": 1,
                }
            )
        rows.append(visible_row)

    return {
        "kind": "spreadsheet",
        "name": file_path.name,
        "sheet_count": 1,
        "preview_limits": {
            "max_rows": preview_row_limit,
            "max_columns": preview_col_limit,
        },
        "sheets": [
            {
                "name": "Sheet1",
                "row_count": row_count,
                "column_count": column_count,
                "preview_row_count": limited_rows,
                "preview_column_count": limited_cols,
                "truncated_rows": row_count > limited_rows,
                "truncated_columns": column_count > limited_cols,
                "column_widths": column_widths,
                "rows": rows,
            }
        ],
    }


def _build_spreadsheet_rows(sheet, limited_rows: int, limited_cols: int, merged_map: dict[tuple[int, int], dict[str, int]]) -> list[list[dict[str, object]]]:
    rows: list[list[dict[str, object]]] = []
    if not limited_rows or not limited_cols:
        return rows

    for row_index in range(1, limited_rows + 1):
        visible_row: list[dict[str, object]] = []
        for col_index in range(1, limited_cols + 1):
            merge = merged_map.get((row_index, col_index))
            if merge and not merge["is_anchor"]:
                continue
            value = sheet.cell(row=row_index, column=col_index).value
            visible_row.append(
                {
                    "value": _coerce_spreadsheet_cell(value),
                    "column": col_index - 1,
                    "rowspan": merge["rowspan"] if merge else 1,
                    "colspan": merge["colspan"] if merge else 1,
                }
            )
        rows.append(visible_row)
    return rows


def _build_merge_map(sheet, limited_rows: int, limited_cols: int) -> dict[tuple[int, int], dict[str, int]]:
    merge_map: dict[tuple[int, int], dict[str, int]] = {}
    if not limited_rows or not limited_cols:
        return merge_map

    merged_ranges = getattr(getattr(sheet, "merged_cells", None), "ranges", None) or []
    for merged_range in merged_ranges:
        min_row = max(1, merged_range.min_row)
        min_col = max(1, merged_range.min_col)
        max_row = min(limited_rows, merged_range.max_row)
        max_col = min(limited_cols, merged_range.max_col)
        if min_row > limited_rows or min_col > limited_cols or max_row < min_row or max_col < min_col:
            continue
        for row_index in range(min_row, max_row + 1):
            for col_index in range(min_col, max_col + 1):
                merge_map[(row_index, col_index)] = {
                    "is_anchor": row_index == min_row and col_index == min_col,
                    "rowspan": max_row - min_row + 1,
                    "colspan": max_col - min_col + 1,
                }
    return merge_map


def _column_letter(index: int) -> str:
    result = ""
    current = index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _excel_width_to_px(width: float | None) -> int:
    if width is None:
        return 96
    return max(72, min(320, int(width * 7 + 12)))


def _infer_text_width_px(text: str) -> int:
    return max(72, min(280, 24 + len(text[:24]) * 8))


def run_demo_server(host: str = "127.0.0.1", port: int = 8080) -> None:
    settings = load_settings(argv=[])
    ensure_chat_history_schema(settings)
    if settings.backend == "docker":
        validate_docker_backend_access(
            container_name=settings.docker_container_name,
            workspace_dir=settings.docker_workspace_dir,
            timeout=min(settings.docker_timeout, 10),
        )

    server = ThreadingHTTPServer((host, port), DemoRequestHandler)
    print(f"Demo server listening on http://{host}:{port}")
    print("POST /api/demo/run 会执行真实 create_deep_agent 调度。")
    if settings.backend == "docker":
        print(
            "Docker backend 已通过启动自检："
            f" container={settings.docker_container_name}, workspace={settings.docker_workspace_dir}"
        )
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
